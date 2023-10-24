#! python3
# aim_output_to_excel_sheet.py converts data from output file aim calculation into excel format

import glob, openpyxl, re
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter

column_titles = (
                 '_CPs_',
                 'Density of all electrons:', 
                 'Lagrangian kinetic energy G(r):', 
                 'Hamiltonian kinetic energy K(r):',
                 'Potential energy density V(r):',
                 'Energy density E(r) or H(r):', 
                 'Laplacian of electron density:',
                 'Electron localization function (ELF):',
                 'Localized orbital locator (LOL):', 
                 'Local information entropy:',
                 'Reduced density gradient (RDG):',
                 'Reduced density gradient with promolecular approximation:',
                 'Sign(lambda2)*rho:',
                 'Sign(lambda2)*rho with promolecular approximation:',
                 'Wavefunction value for orbital',
                 'Average local ionization energy (ALIE):',
                 'Delta_g (under promolecular approximation):',
                 'Delta_g (under Hirshfeld partition):',
                 'User-defined real space function:',
                 'ESP from nuclear charges:',
                 'Norm of gradient is:', # Components of gradient in x/y/z are:
                 'Total:', # Components of Laplacian in x/y/z are:
                 'Determinant of Hessian:',
                 'Ellipticity of electron density:',
                 'eta index:',
                 )


# Create Workbook
wb = openpyxl.Workbook()

# Create sheet name and column titles (+style settings)
sheet = wb.active
sheet.title = 'first'
sheet.freeze_panes = 'B2' # Fixing some areas

for i in range(0, len(column_titles)):
    sheet.cell(row=1, column=i+1).value = column_titles[i]
    sheet.cell(row=1, column=i+1).alignment = Alignment(horizontal='center', 
                                                        vertical='center',
                                                        wrap_text=True)
    sheet.column_dimensions[get_column_letter(i+2)].width = 30
sheet.row_dimensions[1].height = 40


# Search data by word in column titles from the output aim
E_regex = re.compile(r'\W\d*\.\d\S*')

file = open('CPprop.txt', 'r')
x = 1 # row number
y = 2 # column number
for line in iter(file):
    if 'CP   ' in line:
        y = 2
        x += 1
        sheet.cell(row=x, column=1).value = 'CP' + str(x-1)
    for i in column_titles:
        if i in line:
            E_regex_search = E_regex.search(line)
            sheet.cell(row=x, column=y).value = E_regex_search.group()
            y += 1


wb.save('aim_results.xlsx')