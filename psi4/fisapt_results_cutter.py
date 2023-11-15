#! python3
# aim_output_to_excel_sheet.py converts data from output file aim calculation into excel format

import glob, openpyxl, itertools
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter


parameter = (
    'Electrostatics',
    'Elst10,r',
    'Exchange',
    'Exch10',
    'Exch10(S^2)',
    'Induction',
    'Ind20,r',
    'Exch-Ind20,r',
    'delta HF,r (2)',
    'Induction (A<-B)',
    'Induction (B<-A)',
    'Dispersion',
    'Disp20',
    'Exch-Disp20',
    'Total HF',
    'Total SAPT0',
)

wb = openpyxl.Workbook()

sheet = wb.active
sheet.title = 'first'
sheet.freeze_panes = 'B2' # Fixing some areas

for i in range(len(parameter)):
    sheet.cell(row=1, column=i+2).value = parameter[i]
    sheet.cell(row=1, column=i+2).alignment = Alignment(horizontal='center',
                                                          vertical='center',
                                                          wrap_text=True)
    sheet.column_dimensions[get_column_letter(i + 2)].width = 20
sheet.row_dimensions[1].height = 40

r = 0
c = 0
for file in glob.glob('*.out'):
    reading = open(file, 'r')
    sheet.cell(row=2+r, column=1).value = file
    for line in iter(reading):
        if '[kcal/mol]' in line:
            sheet.cell(row=2+r, column=2+c).value = line[53:69]
            c += 1
    r += 1
    c = 0
wb.save('fisapt_results.xlsx')
