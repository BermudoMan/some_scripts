import openpyxl, re

wb = openpyxl.Workbook()
sheet = wb.active

marker = 'data_CSD_CIF_'
cif = 'Au_C_compounds.cif'
i = 0

bond_length_format = re.compile(r'\d.\d\d\d')

'----Search Templates----'
"""Au-Au"""
# bond_re_format = re.compile(r'Au\w* Au\w')
"""Au-P"""
# bond_re_format = re.compile(r'Au\w* P\w|P\w* Au\w')
"""Au-S"""
# bond_re_format = re.compile(r'Au\w* S\w|S\w* Au\w')
"""Au-C"""
bond_re_format = re.compile(r'Au\w*\sC[^lsudnoa]|^C[^lsudnoa]*\sAu\w')

from_ccdc_data = open(cif, 'r')
for line in from_ccdc_data:
    bond_re_format_found = bond_re_format.search(line)
    if marker in line:
        sheet.cell(row=1 + i, column=3).value = line[13:19]
    try:
        bond_re_format_found.group()
        bond_length_format_found = bond_length_format.search(line)
        sheet.cell(row=1 + i, column=1).value = line
        sheet.cell(row=1 + i, column=2).value = bond_length_format_found.group()
        i += 1
    except:
        print('NaN was found in the line')

wb.save(cif.replace('.cif', '') + '.xlsx')
