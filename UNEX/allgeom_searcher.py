#! python3
import pathlib
from pathlib import Path
import openpyxl
from itertools import islice
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter

paths = sorted(Path('.').glob('*.ks'))
file_in_directory = list(map(str, paths))
print(file_in_directory)
wb = openpyxl.Workbook()

start_marker = 'Internal geometrical parameters'
end_marker = '-------------------------------------------------------------------------'
x = 0
for i in file_in_directory:
    file = open(i, 'r')
    sheet = wb.create_sheet(index=0, title=str(i))
    sheet = wb.active
    for index, line in enumerate(file):
        if start_marker in line:
            start_index = index
            print(start_index)
            end_index = start_index+158
    file.seek(0)
    for line in islice(file, start_index+5, end_index+5):
        line_data = line.split()
        if 'dist' in line:
            for i in range(len(line_data)):
                sheet.cell(row=1+x, column=i+1).value = line_data[i]
            sheet['I' + str(line_data[0])] = '=SQRT((2.5*H' + str(line_data[0]) + '^2+(0.002*G' + str(line_data[0]) + (')^2))')
            sheet['J' + str(line_data[0])] = '=TEXT(G' + str(line_data[0]) + ',"0.000")'
            sheet['K' + str(line_data[0])] = '=TEXT(I' + str(line_data[0]) + '*1000,"0")'
            sheet['L' + str(line_data[0])] = '=(J' + str(line_data[0]) + '&"("&K' + str(line_data[0]) + '&")")'
        elif 'angle' in line:
            for i in range(len(line_data)):
                sheet.cell(row=1+x, column=i+2).value = line_data[i]
            sheet['I' + str(line_data[0])] = '=3*H' + str(line_data[0])
            sheet['J' + str(line_data[0])] = '=TEXT(G' + str(line_data[0]) + ',"0.0")'
            sheet['K' + str(line_data[0])] = '=TEXT(I' + str(line_data[0]) + '*10,"0")'
            sheet['L' + str(line_data[0])] = '=(J' + str(line_data[0]) + '&"("&K' + str(line_data[0]) + '&")")'
        elif 'torsion'in line:
            for i in range(len(line_data)):
                sheet.cell(row=1+x, column=i+1).value = line_data[i]
            sheet['I' + str(line_data[0])] = '=3*H' + str(line_data[0])
            sheet['J' + str(line_data[0])] = '=TEXT(G' + str(line_data[0]) + ',"0.0")'
            sheet['K' + str(line_data[0])] = '=TEXT(I' + str(line_data[0]) + '*10,"0")'
            sheet['L' + str(line_data[0])] = '=(J' + str(line_data[0]) + '&"("&K' + str(line_data[0]) + '&")")'
        x += 1
    x = 0

wb.save('unex_allgeom.xlsx')