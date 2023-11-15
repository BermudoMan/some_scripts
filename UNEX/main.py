import pathlib
from pathlib import Path
import openpyxl
from itertools import islice
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter

paths = sorted(Path('.').glob('*.ks'))
file_in_directory = list(map(str, paths))
print(file_in_directory)
wb = openpyxl.Workbook()

start_marker = 'Internal geometrical parameters'
end_marker = '-------------------------------------------------------------------------'
x = 0
for i in file_in_directory:
    file = open(i, 'r')
    sheet = wb.create_sheet(index=0, title=str(i))
    sheet = wb.active
    for index, line in enumerate(file):
        if start_marker in line:
            start_index = index
            print(start_index)
            end_index = start_index+180
    file.seek(0)
    for line in islice(file, start_index, end_index):
        sheet.cell(row=1+x, column=1).value = line
        x += 1
    x = 0

wb.save('unex_allgeom.xlsx')