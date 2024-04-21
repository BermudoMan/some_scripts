#! python3
import pathlib
from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from itertools import islice
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
import re

paths = sorted(Path('.').glob('*.ks'))
file_in_directory = list(map(str, paths))
print(file_in_directory)


# start_marker = 'Internal geometrical parameters'
# end_marker = '-------------------------------------------------------------------------'
# x = 0

def separate_list_data():
    wb = openpyxl.Workbook()

    start_marker = 'Internal geometrical parameters'
    end_marker = '-------------------------------------------------------------------------'
    x = 0
    for i in file_in_directory:
        file = open(i, 'r')
        sheet = wb.create_sheet(index=0, title=str(i))
        sheet = wb.active
        for index, line in enumerate(file):
            if start_marker in line:
                start_index = index
                print(start_index)
                end_index = start_index+158
        file.seek(0)
        for line in islice(file, start_index+5, end_index-23):
            line_data = line.split()
            if 'dist' in line:
                for i in range(len(line_data)):
                    sheet.cell(row=1+x, column=i+1).value = line_data[i]
                sheet['I' + str(line_data[0])] = '=SQRT(( (2.5*H' + str(line_data[0]) + ')^2 + (0.002*G' + str(line_data[0]) + ')^2) )'
                sheet['J' + str(line_data[0])] = '=TEXT(G' + str(line_data[0]) + ',"0.000")'
                sheet['K' + str(line_data[0])] = '=ROUNDUP(I' + str(line_data[0]) + ',"3")'
                sheet['L' + str(line_data[0])] = '=TEXT(K' + str(line_data[0]) + '*1000, "0")'
                sheet['M' + str(line_data[0])] = '=(J' + str(line_data[0]) + '&"("&L' + str(line_data[0]) + '&")")'
            elif 'angle' in line:
                for i in range(len(line_data)):
                    sheet.cell(row=1+x, column=i+2).value = line_data[i]
                sheet['I' + str(line_data[0])] = '=3*H' + str(line_data[0])
                sheet['J' + str(line_data[0])] = '=TEXT(G' + str(line_data[0]) + ',"0.0")'
                sheet['K' + str(line_data[0])] = '=ROUNDUP(I' + str(line_data[0]) + ',"1")'
                sheet['L' + str(line_data[0])] = '=TEXT(I' + str(line_data[0]) + '*10,"0")'
                sheet['M' + str(line_data[0])] = '=(J' + str(line_data[0]) + '&"("&L' + str(line_data[0]) + '&")")'
            elif 'torsion'in line:
                for i in range(len(line_data)):
                    sheet.cell(row=1+x, column=i+1).value = line_data[i]
                sheet['I' + str(line_data[0])] = '=3*H' + str(line_data[0])
                sheet['J' + str(line_data[0])] = '=TEXT(G' + str(line_data[0]) + ',"0.0")'
                sheet['K' + str(line_data[0])] = '=ROUNDUP(I' + str(line_data[0]) + ',"1")'
                sheet['L' + str(line_data[0])] = '=TEXT(I' + str(line_data[0]) + '*10,"0")'
                sheet['M' + str(line_data[0])] = '=(J' + str(line_data[0]) + '&"("&L' + str(line_data[0]) + '&")")'
            elif 'o-o-p'in line:
                for i in range(len(line_data)):
                    sheet.cell(row=1+x, column=i+1).value = line_data[i]
                sheet['I' + str(line_data[0])] = '=3*H' + str(line_data[0])
                sheet['J' + str(line_data[0])] = '=TEXT(G' + str(line_data[0]) + ',"0.0")'
                sheet['K' + str(line_data[0])] = '=ROUNDUP(I' + str(line_data[0]) + ',"1")'
                sheet['L' + str(line_data[0])] = '=TEXT(I' + str(line_data[0]) + '*10,"0")'
                sheet['M' + str(line_data[0])] = '=(J' + str(line_data[0]) + '&"("&L' + str(line_data[0]) + '&")")'
            x += 1
        x = 0
    wb.save('unex_allgeom.xlsx')




def summary_list_data():
    wb = load_workbook('unex_allgeom.xlsx')
    count = 0
    try:
        del wb['Sheet']
    except:
        print('WHAT')
    # names = wb.sheetnames
    names = ['RegAlpha=6.e-5','RegAlpha=7.e-5','RegAlpha=8.e-5','RegAlpha=9.e-5','RegAlpha=1.e-4','RegAlpha=2.e-4','RegAlpha=3.e-4','RegAlpha=4.e-4',
            'RegAlpha=5.e-4','RegAlpha=6.e-4','RegAlpha=7.e-4','RegAlpha=8.e-4','RegAlpha=9.e-4',
            'RegAlpha=1.e-3','RegAlpha=2.e-3','RegAlpha=3.e-3','RegAlpha=4.e-3','RegAlpha=5.e-3',
            'RegAlpha=6.e-3','RegAlpha=7.e-3','RegAlpha=8.e-3','RegAlpha=9.e-3','RegAlpha=1.e-2']

    # del_.active()
    # wb.remove(del_)

    sheet = wb.create_sheet(index=0, title='summary')
    print(names)

    
    re_model = re.compile(r'\d*.\d\d\d')

    for i_ in names:
        i = i_ + '.ks'
        try:
            fr = open('fr.txt', 'r')
            for line_fr in fr:
                if i in line_fr:
                    re_model_search = re_model.search(line_fr)
                    fr_value = re_model_search.group()


            sheet_i = wb[i]

            sheet.cell(row=1, column=count+7).value = i.replace('RegAlpha=', '')
            sheet.cell(row=2, column=count+7).value = fr_value

            for row_num_labels in range(len(sheet_i['L'])):
                sheet.cell(row=(row_num_labels+4), column=1).value = "='" + i + "'!$A" + str(row_num_labels) 
                sheet.cell(row=(row_num_labels+4), column=2).value = "='" + i + "'!$B" + str(row_num_labels) 
                sheet.cell(row=(row_num_labels+4), column=3).value = "='" + i + "'!$C" + str(row_num_labels) 
                sheet.cell(row=(row_num_labels+4), column=4).value = "='" + i + "'!$D" + str(row_num_labels) 
                sheet.cell(row=(row_num_labels+4), column=5).value = "='" + i + "'!$E" + str(row_num_labels) 
                sheet.cell(row=(row_num_labels+4), column=6).value = "='" + i + "'!$F" + str(row_num_labels) 

            for row_num in range(len(sheet_i['L'])):
                sheet.cell(row=(row_num+4), column=7+count).value = "='" + i + "'!$M" + str(row_num) 
                # sheet['B1' + str(row_num)] = '=' + i + '!$L' + str(row_num) 

            count +=1
            fr.close()
        except:
            print('WHAT?')
    wb.save('unex_allgeom.xlsx')

separate_list_data()
summary_list_data()
