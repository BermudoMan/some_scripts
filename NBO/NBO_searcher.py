import glob
from itertools import islice
import pandas as pd
import openpyxl

start = ' Second Order Perturbation Theory Analysis of Fock Matrix in NBO Basis'
stop = ' Natural Bond Orbitals (Summary):'


def counter_x():
    for x, value in enumerate(f):
        if start in value:
            return x + 1


def counter_y():
    f.seek(0)
    for y, value in enumerate(f):
        if stop in value:
            return y + 1



words = [
            'C   1 - C   4', 'C   1 - C   4', 'O   7 - C   9', 'O   8 - C  20', 
            'C   1 - O   7', 'C   9 - C  10', 'C   9 - C  11', 
            'C   4 - O   8', 'C  20 - C  21', 'C  20 - C  22',
            'LP (   1) O   7', 'LP (   2) O   7', 'LP (   1) O   8', 'LP (   2) O   8'    
        ]
for file in glob.glob("*.log"):
    f = open(file, 'r')
    name = file + '.txt'
    NBO = open(name, "+w")

    sort1 = 'RY'
    sort2 = 'CR'
    NBO.write(file)
    NBO.write(2 * "\n")

    x = counter_x()
    print(x)
    y = counter_y()
    print(y)

    for word in words:
        NBO.write(2 * "\n")
        NBO.write("!!!!!!" + word + "!!!!!!")
        NBO.write(2 * "\n")
        for line in islice(f, x, y):
            if word in line:
                if sort1 not in line and sort2 not in line:
                    NBO.write(line)

        f.seek(0)
    f.close()
    NBO.close()




for file in glob.glob("*.txt"):
    open_file = open(file, 'r')
    namer = file + '!'
    result = open(namer, "+w")

    df = pd.DataFrame(open_file)
    result_df = df.drop_duplicates()
    string = result_df.to_string(header=False, index=False)
    result_df.style.set_properties(**{'text-align': 'left'})
    result.write(string)
    # df_name = file.replace('.log.txt', '') + '.xlsx'
    # result_df.to_excel(df_name)
    open_file.close()


wb = openpyxl.Workbook()
cou = 0
for file in glob.glob("*.txt!"):
    f = open(file, 'r')
    nice_name = str(file).replace('.log.txt!', '')
    sheet = wb.create_sheet(index=0, title=nice_name)
    sheet = wb.active
    sheet.column_dimensions['A'].width = 8
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 5
    for line in f:
        # for i in range(len(line_data)):
        sheet.cell(row=1+cou, column=1).value = line
        sheet.cell(row=1+cou, column=2).value = line[16:32]
        sheet.cell(row=1+cou, column=3).value = line[53:70]
        sheet.cell(row=1+cou, column=4).value = line[12:14] + '_' + line[49:51]
        sheet.cell(row=1+cou, column=5).value = line[74:82]
        cou += 1
    cou = 0
wb.save('nbo_summary.xlsx')





